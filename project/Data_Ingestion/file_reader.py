"""
Polars-based ingestion engine for Phase 1.

Supports:
  1. WIDE layout (rows=metrics, columns=periods)
  2. TALL layout (rows=periods, columns=metrics)
  3. LEDGER/TALLY style exports (detected heuristically)
"""

import re
from pathlib import Path
from typing import Optional

import duckdb
import polars as pl

from canonical_fields import (
    resolve_alias,
    detect_ledger_format,
    extract_ledger_facts,
)
from periods import parse_period
from units import normalise


def list_excel_sheets(file_path: Path) -> list[str]:
    """Return sheet names from an Excel file, or [''] for CSV."""
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        return sheets
    return [""]


_DISCLAIMER_RE = re.compile(
    r"all numbers|unless mentioned|figures in|amounts in|values in|"
    r"note:|notes:|source:|currency:|in rs\.|in inr\b",
    re.I,
)

_SKIP_FILENAME_WORDS = {
    "pnl",
    "pl",
    "balance",
    "sheet",
    "bs",
    "monthly",
    "annual",
    "quarterly",
    "audited",
    "accounts",
    "mis",
    "management",
    "report",
    "financials",
    "financial",
    "data",
    "fy",
    "fy22",
    "fy23",
    "fy24",
    "fy25",
    "fy26",
    "cim",
    "deck",
    "model",
    "final",
    "v1",
    "v2",
    "v3",
    "revised",
    "draft",
    "copy",
    "export",
    "tally",
    "summary",
    "detail",
    "detailed",
    "overview",
    "income",
    "statement",
    "profit",
    "loss",
    "p&l",
}


def _find_data_start(file_path: Path, sheet_name: Optional[str] = None) -> tuple[int, list[str]]:
    """
    Scan first rows to find the effective header row and preamble lines.
    Strategy: pick the first row close to max non-empty cells.
    """
    import csv as _csv

    suffix = file_path.suffix.lower()
    rows: list[list[str]] = []

    if suffix == ".csv":
        with open(file_path, encoding="utf-8-sig", errors="replace") as f:
            reader = _csv.reader(f)
            for _, row in zip(range(30), reader):
                rows.append([str(v).strip() for v in row])
    else:
        if sheet_name:
            _sheet: str | int = sheet_name
        else:
            import openpyxl as _openpyxl

            _wb = _openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            _sheet = _wb.sheetnames[0]
            _wb.close()

        raw = pl.read_excel(file_path, sheet_name=_sheet, infer_schema_length=0, has_header=False)
        for row in raw.head(30).iter_rows():
            rows.append([str(v).strip() if v is not None else "" for v in row])

    if not rows:
        return 0, []

    counts = [sum(1 for v in row if v) for row in rows]
    max_count = max(counts, default=0)
    if max_count == 0:
        return 0, []

    header_idx = next((i for i, c in enumerate(counts) if c >= max(2, max_count - 1)), 0)

    preamble: list[str] = []
    for i in range(header_idx):
        non_empty = [v for v in rows[i] if v]
        if len(non_empty) == 1:
            preamble.append(non_empty[0])

    return header_idx, preamble


def _entity_name_from_filename(file_path: Path) -> str:
    words = re.split(r"[_\-\s]+", file_path.stem)
    kept = [
        w.title()
        for w in words
        if not re.match(r"^\d{2,4}$", w) and w.lower() not in _SKIP_FILENAME_WORDS
    ]
    return " ".join(kept) if kept else file_path.stem.replace("_", " ").title()


def _entity_name_from_preamble(preamble: list[str]) -> Optional[str]:
    for line in preamble:
        if _DISCLAIMER_RE.search(line):
            continue
        name = re.sub(r"\(.*?\)", "", line).strip()
        name = re.split(r"\s*[-|:]\s*", name)[0].strip()
        words = re.split(r"[\s_]+", name)
        kept = [
            w
            for w in words
            if w and not re.match(r"^\d{2,4}$", w) and w.lower() not in _SKIP_FILENAME_WORDS
        ]
        result = " ".join(kept)
        if len(result) >= 3:
            return result
    return None


def detect_entity_name(file_path: Path, df: pl.DataFrame, preamble: Optional[list[str]] = None) -> str:
    if preamble:
        from_preamble = _entity_name_from_preamble(preamble)
        if from_preamble:
            return from_preamble
    return _entity_name_from_filename(file_path)


PERIOD_PATTERN = re.compile(
    r"(FY\s*\d{2,4}|Q[1-4]\s*FY\s*\d{2,4}|FY\s*\d{2,4}[\-_]Q[1-4]"
    r"|FY\s*\d{2,4}[\-_]M\d{1,2}"
    r"|(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s\-]\d{2,4}"
    r"|\d{4})",
    re.I,
)


def _looks_like_period(val: str) -> bool:
    return bool(PERIOD_PATTERN.match(str(val).strip()))


def detect_layout(df: pl.DataFrame) -> str:
    period_cols = sum(1 for c in df.columns if _looks_like_period(c))
    if period_cols >= 2:
        return "wide"

    first_col = df.columns[0]
    sample = df[first_col].drop_nulls().head(10).cast(pl.Utf8).to_list()
    period_rows = sum(1 for v in sample if _looks_like_period(v))
    if period_rows >= 2:
        return "tall"

    return "wide"


def _read_file(file_path: Path, sheet_name: Optional[str] = None) -> tuple[pl.DataFrame, str, list[str], int]:
    suffix = file_path.suffix.lower()
    skip_rows, preamble = _find_data_start(file_path, sheet_name)

    if suffix in (".xlsx", ".xls"):
        if sheet_name:
            _sheet: str | int = sheet_name
        else:
            import openpyxl as _openpyxl

            _wb = _openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            _sheet = _wb.sheetnames[0]
            _wb.close()

        df_raw = pl.read_excel(file_path, sheet_name=_sheet, infer_schema_length=0, has_header=False)
        if len(df_raw) > skip_rows:
            header_vals = [
                str(v) if v is not None and str(v) not in ("None", "") else ""
                for v in df_raw.row(skip_rows)
            ]
            seen: dict[str, int] = {}
            unique_headers: list[str] = []
            for h in header_vals:
                if not h:
                    h = "__UNNAMED__"
                if h in seen:
                    seen[h] += 1
                    unique_headers.append(f"{h}__{seen[h]}")
                else:
                    seen[h] = 0
                    unique_headers.append(h)
            df = df_raw.slice(skip_rows + 1).rename({old: new for old, new in zip(df_raw.columns, unique_headers)})
        else:
            df = df_raw
    elif suffix == ".csv":
        df = pl.read_csv(file_path, infer_schema_length=0, truncate_ragged_lines=True, skip_rows=skip_rows)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    detected_unit = _sniff_unit(df, preamble)
    return df, detected_unit, preamble, skip_rows


def _sniff_unit(df: pl.DataFrame, preamble: Optional[list[str]] = None) -> str:
    unit_hints = [
        "crore",
        "lakh",
        "lacs",
        "lakhs",
        "cr.",
        "rs.",
        "inr",
        "usd",
        "thousand",
        "000",
        "million",
        "rupee",
    ]
    scale_hints = ["crore", "lakh", "lacs", "lakhs", "cr.", "thousand", "000", "million"]

    candidates: list[str] = list(preamble or []) + list(df.columns)
    if len(df) > 0:
        col0 = df.columns[0]
        candidates.extend(df[col0].head(3).cast(pl.Utf8).to_list())

    for candidate in candidates:
        s = str(candidate).lower()
        if any(h in s for h in scale_hints):
            return str(candidate)
    for candidate in candidates:
        s = str(candidate).lower()
        if any(h in s for h in unit_hints):
            return str(candidate)
    return ""


def _parse_numeric_value(raw_val) -> Optional[float]:
    if raw_val is None:
        return None

    val_str = str(raw_val).strip()
    if not val_str or val_str in ("", "-", "—", "N/A", "na", "n/a"):
        return None

    if val_str.startswith("(") and val_str.endswith(")"):
        val_str = "-" + val_str[1:-1]

    val_str = val_str.replace(",", "").strip()

    try:
        return float(val_str)
    except ValueError:
        return None


_ROW_UNIT_RE = re.compile(
    r"\b(rs\.?|inr|usd)?\s*(crore|crores|cr\.?|lakh|lakhs|lacs|million|mn|thousand|000)\b",
    re.I,
)
_ABS_HINT_RE = re.compile(r"\b(rs\.?|inr|usd)\s*(absolute|actual|exact|value|amount)\b", re.I)


def _row_level_unit_override(label: str, fallback_unit: str = "") -> Optional[str]:
    label = label or ""
    ll = label.lower()

    if _ABS_HINT_RE.search(label):
        if "usd" in ll:
            return "USD absolute"
        return "INR absolute"

    # Handle parenthetical unit hints in row names: e.g. "Packaging (Rs absolute)"
    paren = re.search(r"\(([^)]*)\)", label)
    if paren:
        inner = paren.group(1)
        m_inner = _ROW_UNIT_RE.search(inner)
        if m_inner:
            token = m_inner.group(0).lower()
            if "usd" in token:
                return f"USD {token}"
            return f"INR {token}"

    m = _ROW_UNIT_RE.search(label)
    if m:
        token = m.group(0).lower()
        if "usd" in token:
            return f"USD {token}"
        return f"INR {token}"

    # Heuristic for mixed-unit sheets: if fallback is lakh/crore and row label explicitly says rupees,
    # treat it as absolute.
    if "rupee" in ll or re.search(r"\brs\b", ll):
        if any(x in (fallback_unit or "").lower() for x in ["lakh", "lac", "crore", "mn", "million"]):
            if "usd" in (fallback_unit or "").lower():
                return "USD absolute"
            return "INR absolute"

    return None


def _get_cell_reference(col_idx: int, row_idx: int) -> str:
    col_letter = ""
    col_num = col_idx
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return f"{col_letter}{row_idx}"


def extract_wide(
    df: pl.DataFrame,
    file_id: int,
    entity_id: int,
    unit_spec_str: str,
    skip_rows: int = 0,
    source_sheet: Optional[str] = None,
) -> list[dict]:
    facts = []
    metric_col = df.columns[0]
    period_cols = [c for c in df.columns[1:] if _looks_like_period(c)]

    if not period_cols:
        period_cols = df.columns[1:]

    for row_idx, row in enumerate(df.iter_rows(named=True)):
        raw_label = str(row.get(metric_col, "") or "").strip()
        if not raw_label:
            continue

        if raw_label.startswith("---"):
            continue

        if re.match(r"^\s*(total|grand total|subtotal|sub[\s\-]total)\s*$", raw_label, re.I):
            if not resolve_alias(raw_label):
                continue

        canonical = resolve_alias(raw_label)
        if not canonical:
            continue

        for col_idx, period_col in enumerate(period_cols):
            period_spec = parse_period(str(period_col))
            if not period_spec:
                continue

            raw_val = row.get(period_col)
            numeric_val = _parse_numeric_value(raw_val)
            if numeric_val is None:
                continue

            row_unit = _row_level_unit_override(raw_label, fallback_unit=unit_spec_str)
            effective_unit = row_unit if row_unit else unit_spec_str
            nv = normalise(numeric_val, effective_unit)
            cell_ref = _get_cell_reference(col_idx + 2, row_idx + 1 + skip_rows)

            facts.append(
                {
                    "file_id": file_id,
                    "entity_id": entity_id,
                    "canonical_field": canonical,
                    "period": period_spec.canonical,
                    "value_normalised": nv.value_normalised,
                    "currency": nv.currency,
                    "original_unit": nv.original_unit,
                    "conversion_factor": nv.conversion_factor,
                    "conversion_applied": nv.conversion_applied,
                    "raw_value": numeric_val,
                    "raw_header": raw_label,
                    "row_context": raw_label,
                    "cell_reference": cell_ref,
                    "source_sheet": source_sheet,
                }
            )

    return facts


def extract_tall(
    df: pl.DataFrame,
    file_id: int,
    entity_id: int,
    unit_spec_str: str,
    skip_rows: int = 0,
    source_sheet: Optional[str] = None,
) -> list[dict]:
    facts = []
    period_col = df.columns[0]
    metric_cols = df.columns[1:]

    for row_idx, row in enumerate(df.iter_rows(named=True)):
        raw_period = str(row.get(period_col, "") or "").strip()
        if raw_period.startswith("---"):
            continue
        period_spec = parse_period(raw_period)
        if not period_spec:
            continue

        for col_idx, metric_col in enumerate(metric_cols, start=1):
            canonical = resolve_alias(metric_col)
            if not canonical:
                continue

            raw_val = row.get(metric_col)
            numeric_val = _parse_numeric_value(raw_val)
            if numeric_val is None:
                continue

            row_unit = _row_level_unit_override(metric_col, fallback_unit=unit_spec_str)
            effective_unit = row_unit if row_unit else unit_spec_str
            nv = normalise(numeric_val, effective_unit)
            cell_ref = _get_cell_reference(col_idx + 1, row_idx + 1 + skip_rows)
            facts.append(
                {
                    "file_id": file_id,
                    "entity_id": entity_id,
                    "canonical_field": canonical,
                    "period": period_spec.canonical,
                    "value_normalised": nv.value_normalised,
                    "currency": nv.currency,
                    "original_unit": nv.original_unit,
                    "conversion_factor": nv.conversion_factor,
                    "conversion_applied": nv.conversion_applied,
                    "raw_value": numeric_val,
                    "raw_header": metric_col,
                    "row_context": raw_period,
                    "cell_reference": cell_ref,
                    "source_sheet": source_sheet,
                }
            )

    return facts


def write_to_staging(conn: duckdb.DuckDBPyConnection, facts: list[dict]) -> int:
    if not facts:
        return 0

    conn.execute("""
        CREATE SEQUENCE IF NOT EXISTS staging_id_seq START 1
    """)

    for f in facts:
        conn.execute(
            """
            INSERT INTO staging_facts (
                staging_id, file_id, entity_id, canonical_field, period,
                value_normalised, currency, original_unit, conversion_factor,
                conversion_applied, raw_value, raw_header, row_context,
                cell_reference, source_sheet
            ) VALUES (
                nextval('staging_id_seq'), ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?
            )
            """,
            [
                f["file_id"],
                f["entity_id"],
                f["canonical_field"],
                f["period"],
                f["value_normalised"],
                f["currency"],
                f["original_unit"],
                f["conversion_factor"],
                f["conversion_applied"],
                f["raw_value"],
                f["raw_header"],
                f["row_context"],
                f.get("cell_reference"),
                f.get("source_sheet"),
            ],
        )

    conn.commit()
    return len(facts)


_ANN_COL_RE = re.compile(r"^(note|annotation|remark|comment|footnote|addendum)\s*", re.I)


def _detect_numerical_claims_from_text(text: str) -> bool:
    return bool(re.search(r"\d", text))


def extract_inline_annotations(df: pl.DataFrame, file_id: int, entity_id: int) -> list[dict]:
    annotation_chunks: list[dict] = []

    for col in df.columns:
        if not _ANN_COL_RE.match(col):
            continue
        for row_idx, val in enumerate(df[col].to_list()):
            text = str(val).strip()
            if not text or text in ("N/A", "n/a", "-", ""):
                continue
            annotation_chunks.append(
                {
                    "file_id": file_id,
                    "entity_id": entity_id,
                    "chunk_index": row_idx,
                    "region_type": "inline_annotation",
                    "chunk_type": "footnote",
                    "section_path": col,
                    "linked_fact_ids": "[]",
                    "linked_periods": "[]",
                    "linked_metrics": "[]",
                    "contains_numerical_claim": _detect_numerical_claims_from_text(text),
                    "numerical_claims": "[]",
                    "raw_text": text,
                    "chroma_document_id": None,
                }
            )

    return annotation_chunks


def ingest_file(
    conn: duckdb.DuckDBPyConnection,
    file_path: Path,
    file_id: int,
    entity_id: int,
    confirmed_unit: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> dict:
    """Full ingestion pipeline for one file/sheet."""
    df, detected_unit, preamble, skip_rows = _read_file(file_path, sheet_name)
    unit_str = confirmed_unit if confirmed_unit else detected_unit
    layout = detect_layout(df)

    # Detect and route ledger/tally style sheets.
    if detect_ledger_format(df, list(df.columns)):
        layout = "ledger"

    entity_name = detect_entity_name(file_path, df, preamble)

    if layout == "wide":
        facts = extract_wide(df, file_id, entity_id, unit_str, skip_rows=skip_rows, source_sheet=sheet_name)
    elif layout == "tall":
        facts = extract_tall(df, file_id, entity_id, unit_str, skip_rows=skip_rows, source_sheet=sheet_name)
    else:
        facts = extract_ledger_facts(df, file_id, entity_id, unit_str)
        for idx, fact in enumerate(facts, start=1):
            fact.setdefault("cell_reference", _get_cell_reference(2, idx + skip_rows))
            fact.setdefault("source_sheet", sheet_name)

    n = write_to_staging(conn, facts)
    unmapped = _find_unmapped(df, layout)

    # Phase 2 bridge: inline annotation columns to qualitative_chunks.
    try:
        annotations = extract_inline_annotations(df, file_id, entity_id)
        if annotations:
            for ann in annotations:
                conn.execute(
                    """
                    INSERT INTO qualitative_chunks (
                        chunk_id, file_id, entity_id, chunk_index, region_type,
                        chunk_type, section_path, linked_fact_ids, linked_periods,
                        linked_metrics, contains_numerical_claim, numerical_claims,
                        raw_text, chroma_document_id
                    ) VALUES (
                        nextval('staging_id_seq'), ?, ?, ?, ?,
                        ?, ?, ?, ?,
                        ?, ?, ?, ?, ?
                    )
                    """,
                    [
                        ann["file_id"],
                        ann["entity_id"],
                        ann["chunk_index"],
                        ann["region_type"],
                        ann["chunk_type"],
                        ann["section_path"],
                        "[]",
                        "[]",
                        "[]",
                        ann["contains_numerical_claim"],
                        "[]",
                        ann["raw_text"],
                        None,
                    ],
                )
            conn.commit()
    except Exception:
        # qualitative_chunks may not exist in some standalone contexts.
        pass

    return {
        "file_id": file_id,
        "layout": layout,
        "detected_unit": detected_unit,
        "unit_used": unit_str,
        "entity_name_detected": entity_name,
        "facts_staged": n,
        "rows_processed": len(df),
        "unmapped_headers": unmapped,
    }


def _find_unmapped(df: pl.DataFrame, layout: str) -> list[str]:
    if layout in ("wide", "ledger"):
        metric_col = df.columns[0]
        labels = df[metric_col].drop_nulls().cast(pl.Utf8).to_list()
        return [l for l in labels if l and not resolve_alias(l)]

    metric_cols = df.columns[1:]
    return [c for c in metric_cols if not resolve_alias(c) and not _looks_like_period(c)]
